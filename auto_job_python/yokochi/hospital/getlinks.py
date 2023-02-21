import os, time, requests, urllib, math, re
from bs4 import BeautifulSoup
import openpyxl as exel

base_dir = os.path.dirname(__file__)

book = exel.Workbook()
sheet = book.active

HOKKAIDO = 1
OKINAWA  = 47

row = 1

def get_all_hospitals_in_japan():
    """
        都道府県コードをrange関数で生成し、
        都道府県のすべてのデータを取り出す
        get_all_hospitals_in_pref関数を
        呼び出す。
    """
    for pref_code in range(HOKKAIDO, OKINAWA +1):
        get_all_hospitals_in_pref(pref_code)

def get_all_hospitals_in_pref(pref_code):
    """
    医療機関の数をfind_number_of_hospitals関数で求め、
    引数で受け取った都道府県コードを使用しURLを生成。
    そのURLをget_hospital_mainに渡す。
    Args:
        pref_code (_string_): _都道府県コード_
    """
    global row
    target_url = f'https://caremap.jp/facility_medicals/map/page:1/prefIds:{pref_code}/key1:0/list_type:list'
    nb_pages = math.ceil(find_number_of_hospitals(target_url) / 30)
    # 特定の都道府県の1ページに相当するURL
    # for page in range(1, nb_pages + 1):
    for page in range(1, 3):
        target_url = f'https://caremap.jp/facility_medicals/map/page:{page}/prefIds:{pref_code}/key1:0/list_type:list'
        base_url = target_url.split('/map/')[0]
        get_hospital_main(target_url, base_url, pref_code)
    book.save(os.path.join(base_dir, f'hospital.xlsx'))

def get_hospital_main(target_url, base_url, pref_code):
    """
    (1) 引数で受け取った、target_url, base_urlを
      get_hospital_list関数に渡し、最大30件の
      医療機関のurlを取得する。
    (2) write_hospital_details関数を呼び出し、
      1URL分をExcelの1行に書き込む。
    Args:
        target_url (_string_): _医療機関を求めるための元となるURL_
        base_url (_string_): _医療機関コードを含んだパス_
        pref_code (_string_): _都道府県コード_
    """
    global row
    urls = get_hospital_list(target_url, base_url)
    for url in urls[:2]:
        write_hospital_details(url, pref_code)
        print(f"row=${row}")
        time.sleep(30)
        row += 1

def get_hospital_list(target_url, base_url):
    """
    医療機関のページに遷移するためのURLをリスト形式で返す
    """
    html = requests.get(target_url).text
    soup = BeautifulSoup(html, 'html5lib')
    trs = soup.select('#dataList tr[class!="t-center"]')

    urls = []
    for tr in trs:
        a = tr.select('td:nth-child(2) > a')
        urls.append(urllib.parse.urljoin(base_url, a[0]['href']))
    
    return urls

def write_hospital_details(url, pref_code):
    """
    1医療機関の情報をExcelシートの1行に書き込む
    """
    global row
    html = requests.get(url).text
    soup = BeautifulSoup(html, 'html5lib')
    trs = soup.select('table tr')
    dict1 = {}
    dict1['都道府県コード'] = pref_code
    header = []
    for tr in trs:
        ths = tr.select('th')
        tds = tr.select('td')
        if len(tds) == 0 and len(ths) == 2: continue
        if len(tds) == 1:
            dict1[ths[0].text] = tds[0].text 
        if len(ths) == 3:
            header = ths
            continue
        if len(tds) == 2 and len(ths) == 1:
            dict1[f"{ths[0].text}_{header[1].text}"] = tds[0].text
            dict1[f"{ths[0].text}_{header[2].text}"] = tds[1].text
    if row ==1:
        write_hospital_header(dict1)
    
    col = 1
    for _, val in dict1.items():
        c = sheet.cell(row, col)
        c.value = val
        col += 1

def write_hospital_header(dic):
    """
    Excelシートの1行目に見出しを書きこむ
    (1回きりの処理)
    """
    global row
    col = 1
    for header_item, _ in dic.items():
        c = sheet.cell(row, col)
        c.value = header_item
        col += 1
        
    row += 1

def extract_int(s):
    """
    都道府県のページの「抽出件数」のデータから
    「,」や「件」などを取り除き整数に変換して返す。
    Args:
        s (_string_): _「抽出件数」見出しの右側のデータ_

    Returns:
        _int_: _抽出件数を整数型に変換した値_
    """
    if "," in s:
        pattern = "^(([0-9]{1,3}),)+([0-9]{1,3})件$"
        ptn = re.compile(pattern)
        num = int(ptn.sub(r"\g<2>" + r"\g<3>", s))
    else:
        pattern = "^([0-9]{1,3})件$"
        ptn = re.compile(pattern)
        num = int(ptn.sub(r"\g<1>", s))
    return num

def find_number_of_hospitals(url):
    """
    都道府県のページの「抽出件数」の右側のデータを
    取り出し返す。
    """
    html = requests.get(url).text
    soup = BeautifulSoup(html, 'html5lib')
    # soup.selectした結果は「1,234件」のような文字列なので、extract_intで整数値に変換する
    return extract_int(soup.select('table[id!="dataList"] tr:nth-child(5) > td')[0].text)

get_all_hospitals_in_japan()