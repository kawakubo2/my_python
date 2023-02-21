import os, time, requests, urllib, re
from bs4 import BeautifulSoup
import openpyxl as exel

pref_code = 40
base_dir = os.path.join(os.path.dirname(__file__), 'excel')

book = exel.Workbook()
sheet = book.active

# 特定の都道府県の1ページに相当するURL
target_url = 'https://caremap.jp/facility_medicals/map/prefIds:40/key1:0/list_type:list/'

base_url = target_url.split('/map/')[0]
row = 1

def extract_int(s):
    if "," in s:
        pattern = "^(([0-9]{1,3}),)+([0-9]{1,3})件$"
        ptn = re.compile(pattern)
        num = int(ptn.sub(r"\g<2>" + r"\g<3>", s))
    else:
        pattern = "^([0-9]{1,3})件$"
        ptn = re.compile(pattern)
        num = int(ptn.sub(r"\g<1>", s))
    return num

def find_number_of_hospital(url):
    html = requests.get(url).text
    soup = BeautifulSoup(html, 'html5lib')
    # soup.selectした結果は「1,234件」のような文字列なので、extract_intで整数値に変換する
    return extract_int(soup.select('table[id!="dataList"] tr:nth-child(5) > td')[0].text)

def get_hospital_main():
    global row
    urls = get_hospital_list()
    if row == 1:
        write_header(urls[0], row)
    row += 1
    for url in urls[:2]:
        write_hospital_details(url, row)    
        time.sleep(30)
        row += 1

def get_hospital_list():
    """
    医療機関のページに遷移するためのURLをリスト形式で返す
    """
    html = requests.get(target_url).text
    soup = BeautifulSoup(html, 'html5lib')
    trs = soup.select('#dataList tr[class!="t-center"]')
    print(f'#dataList tr:{len(trs)}' )

    urls = []
    for tr in trs:
        a = tr.select('td:nth-child(2) > a')
        urls.append(urllib.parse.urljoin(base_url, a[0]['href']))
    
    return urls

def write_hospital_details(url, row):
    """
    1医療機関の情報をExcelシートの1行に書き込む
    """
    html = requests.get(url).text
    soup = BeautifulSoup(html, 'html5lib')
    trs = soup.select('table tr')
    dict1 = {}
    three_col_header = []
    for tr in trs:
        ths = tr.select('th')
        tds = tr.select('td')
        if len(tds) == 0 and len(ths) == 2: continue
        if len(tds) == 1 and len(ths) == 1:
            dict1[ths[0].text] = tds[0].text 
        if len(tds) == 0 and len(ths) == 3:
            three_col_header = ths
            continue
        if len(tds) == 2 and len(ths) == 1:
            dict1[f"{ths[0].text}_{three_col_header[1].text}"] = tds[0].text
            dict1[f"{ths[0].text}_{three_col_header[2].text}"] = tds[1].text
    
    col = 1
    for _, val in dict1.items():
        c = sheet.cell(row, col)
        c.value = val
        col += 1

def write_header(url, row):
    """
    Excelのヘッダに書きこむためのリストを返す。
    """
    header = []
    html = requests.get(url).text
    soup = BeautifulSoup(html, 'html5lib')
    trs = soup.select('table tr')
    three_col_header = []
    for tr in trs:
        ths = tr.select('th')
        tds = tr.select('td')
        if len(tds) == 0 and len(ths) == 2: continue
        if len(tds) == 1 and len(ths) == 1:
            header.append(ths[0].text)
        if len(tds) == 0 and len(ths) == 3:
            three_col_header = ths
            continue
        if len(tds) == 2 and len(ths) == 1:
            header.append(f"{ths[0].text}_{three_col_header[1].text}")
            header.append(f"{ths[0].text}_{three_col_header[2].text}")

    col = 1
    for h in header:
        c = sheet.cell(row, col)
        c.value = h
        col += 1

if __name__ == '__main__':
    print(find_number_of_hospital(target_url))
    get_hospital_main()
    book.save(os.path.join(base_dir, f'hospital_{pref_code}.xlsx'))