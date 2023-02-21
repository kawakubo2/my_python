import openpyxl as excel, json
import os, sys

class Invoice:
    
    def __init__(self, source, dest_dir, template, subject):
        base_dir = os.path.dirname(__file__)
        self.in_file = os.path.join(base_dir, source)
        self.out_dir = os.path.join(base_dir, dest_dir)
        if not os.path.exists(self.out_dir):
            os.mkdir(self.out_dir)
        else:
            answer = input(f"{out_dir}フォルダは既に存在します。上書きしますか？(yes/no): ")
            if answer == 'no':
                sys.exit()
        self.template_file = os.path.join(base_dir, template)
        self.subject = subject

    # メイン処理
    def gen_invoice(self):
        # JSONファイルを読み込む
        with open(self.in_file, 'rt') as fp:
            users = json.load(fp)
        # 顧客ごとに請求書を作成
        for name, data in users.items():
            self.make_user_invoice(name, data)
            
    def make_user_invoice(self, name, data):
        # テンプレートを読み込む
        book = excel.load_workbook(self.template_file)
        sheet = book.active
        # シートに名前と件名と金額を書き込む
        sheet["B4"] = name
        sheet["C10"] = self.subject
        sheet["C11"] = data["total"]
        # 内訳を連続で書き込む
        for i, it in enumerate(data['items']):
            date, summary, cnt, price = it
            row = i + 15
            sheet.cell(row, 2, f"{summary}({date})")
            sheet.cell(row, 5, cnt)
            sheet.cell(row, 6, price)
            sheet.cell(row, 7, cnt*price)
        # 請求書をファイルに保存
        out_file = os.path.join(self.out_dir, f"{name}様.xlsx")
        book.save(out_file)
        
if __name__ == '__main__':
    import datetime
    invoice_date = datetime.datetime.now()
    out_dir = f"{invoice_date.year}年{invoice_date.month}年請求分"
    subject = f"{invoice_date.month}年のご請求"
    invoice = Invoice('matome.json', out_dir, 'invoice-template.xlsx', subject)
    invoice.gen_invoice()
