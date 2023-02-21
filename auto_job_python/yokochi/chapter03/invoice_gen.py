import openpyxl as excel
import os

base_dir = os.path.dirname(__file__)
template_file = 'invoice-template.xlsx'
save_file = 'invoice01.xlsx'

name='田中一郎'
subject='1月分のご請求'
items=[
    ['リンゴ', 5, 320],
    ['バナナ', 8, 210],
    ['メロン', 1, 1200]
]

book = excel.load_workbook(os.path.join(base_dir, template_file))
sheet = book.active

sheet['B4'] = name
sheet['C10'] = subject
total = 0
for i, it in enumerate(items):
    summary, count, price = it
    subtotal = count * price
    total += subtotal
    
    row = 15 + i
    sheet.cell(row, 2, summary)
    sheet.cell(row, 5, count)
    sheet.cell(row, 6, price)
    sheet.cell(row, 7, subtotal)
sheet['C11'] = total

book.save(os.path.join(base_dir, save_file))
