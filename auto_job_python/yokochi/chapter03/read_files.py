import openpyxl as excel
import os, glob, pprint

base_dir = os.path.dirname(__file__)

target_dir = 'salesbooks'
save_file = 'matome.xlsx'

def read_files():
    book = excel.Workbook()
    main_sheet = book.active
    # ファイルを列挙して読み込む
    enumfiles(main_sheet)
    # 読み込んだデータを保存
    book.save(os.path.join(base_dir, save_file))
    
def enumfiles(main_sheet):
    # Excelファイルの一覧を得る
    files = glob.glob(os.path.join(base_dir, target_dir, '*.xlsx'))
    # 各Excelブックを次々と読んでいく
    for fname in files:
        read_book(main_sheet, fname)
        
# ブックを開いて中身を読む
def read_book(main_sheet, fname):
    print("read:", fname)
    # Excelブックを読み込む
    book = excel.load_workbook(os.path.join(base_dir, fname), data_only=True)
    sheet = book.active
    # 売上データのある範囲を読み取る
    for row in sheet.iter_rows(min_row=4):
        values = [cell.value for cell in row]
        if values[0] is None: break
        print(values)
        main_sheet.append(values)
        
if __name__ == '__main__':
    read_files()