from email.mime import base
import openpyxl as excel
import os

base_dir = os.path.dirname(__file__)
in_file = os.path.join(base_dir, 'name_combine2.xlsx')
out_file = os.path.join(base_dir, 'name_split2.xlsx')

in_book = excel.load_workbook(in_file)
in_sheet = in_book.worksheets[0]
out_book = excel.Workbook()
out_sheet = out_book.active

for row in in_sheet.iter_rows():
    name = row[0].value
    lastname_len = row[1].value
    lastname = name[:lastname_len]
    firstname = name[lastname_len:]
    out_sheet.append([lastname, firstname])
    
out_book.save(out_file)