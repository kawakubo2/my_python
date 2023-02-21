import openpyxl as excel
import os

base_dir = os.path.dirname(__file__)

in_file = os.path.join(base_dir, 'company.xlsx')
out_file = os.path.join(base_dir, 'company_combine.xlsx')

in_book = excel.load_workbook(in_file)
in_sheet = in_book.worksheets[0]
out_book = excel.Workbook()
out_sheet = out_book.active
for row in in_sheet.iter_rows():
    company_name = row[0].value
    houjin_kaku = row[1].value
    before_after = row[2].value
    name = houjin_kaku + company_name if before_after == '前' else company_name + houjin_kaku
    out_sheet.append([name])
    
out_book.save(out_file)    
    