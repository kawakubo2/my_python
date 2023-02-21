import openpyxl as excel
import os

base_dir = os.path.dirname(__file__)
in_file = os.path.join(base_dir, 'name2.xlsx')
out_file = os.path.join(base_dir, 'name_combine2.xlsx')

in_book = excel.load_workbook(in_file)
in_sheet = in_book.worksheets[0]
out_book = excel.Workbook()
out_sheet = out_book.active
# シートを全部読む
for row in in_sheet.iter_rows():
    sei = row[0].value
    na  = row[1].value
    # 姓と名をつなげる
    name = sei + na
    out_sheet.append([name, len(sei)])

out_book.save(out_file)