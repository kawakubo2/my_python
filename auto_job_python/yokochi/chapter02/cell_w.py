import openpyxl as excel
import os

book = excel.Workbook()
sheet = book.active

# (1)セル名の指定
sheet["A1"] = "勤勉な人の計画は必ず成功する"

# (2)cellメソッドで行番号、列番号、値の指定
sheet.cell(row=2,column=1,value="猿の尻笑い")

# (3)cellオブジェクトを行番号、列番号を指定して取り出し、valueプロパティに値を設定
cell = sheet.cell(row=3, column=1)
cell.value = "探すのに時ありあきらめるのに時がある"

book.save(os.path.dirname(__file__) + '/cell_w.xlsx')

