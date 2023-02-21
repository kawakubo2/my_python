# -*- coding: utf-8 -*-

import openpyxl as excel
import os

book = excel.load_workbook(os.path.dirname(__file__) + '/test100.xlsx')
sheet = book.active

print("セル名から行番号・列番号を取得")
cell = sheet["C2"]
(row, col) = (cell.row, cell.column)
print(f"c2 ---> ({row},{col})")

print("行番号・列番号からセル名を取得")
cell = sheet.cell(row=2, column=3)
coordinate = cell.coordinate
print(f"(2,3) ---> {coordinate}")

def cellname_to_rowcol(sheet, cellname):
    cell = sheet[cellname]
    return (cell.row, cell.column)

def rowcol_to_cellname(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    return cell.coordinate
    
print(cellname_to_rowcol(sheet, "F10"))
print(rowcol_to_cellname(sheet, 10, 6))