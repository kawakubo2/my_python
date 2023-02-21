import openpyxl as excel
import os

book = excel.load_workbook(os.path.dirname(__file__) + '/abc.xlsx')
for sheetname in book.sheetnames:
    print(f"シート名: {sheetname}")
    sheet = book[sheetname]
    print(sheet['A1'].value)

for i in range(len(book.sheetnames)):
    sheet = book.worksheets[i]
    print(sheet['A1'].value)

"""
sheet = book.create_sheet(title='第2四半期売上計画')
sheet['A1'] = '4番目のシート'
sheet = book.create_sheet(title='4月売上計画')
sheet['A1'] = '5番目のシート'
sheet = book.create_sheet(title='4月売上実績')
sheet['A1'] = '6番目のシート'

sheet = book.copy_worksheet(book['第2四半期売上計画'])
sheet.title = '第3四半期売上計画'
sheet['A1'] = '7番目のシート'
"""
book.remove(book['第3四半期売上計画'])


book.save(os.path.dirname(__file__) + '/abc.xlsx')

book.close()