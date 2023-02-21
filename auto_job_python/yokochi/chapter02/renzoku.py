import openpyxl as excel
import os

def fibonacci(n):
    if n in (0, 1):
        return n
    return fibonacci(n - 2) + fibonacci(n - 1)  

book = excel.Workbook()
sheet = book.active
for i in range(20):
    sheet.cell(row=i+1, column=1, value=i)
    sheet.cell(row=i+1, column=2, value=fibonacci(i))

book.save(os.path.dirname(__file__) + '/renzoku.xlsx')