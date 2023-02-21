import openpyxl as excel
import os

book = excel.load_workbook(os.path.dirname(__file__) + '/all-customer.xlsx')
sheet = book['名簿']

plan_dict = {}
for row in sheet.iter_rows(min_row=3):
    cells = [v.value for v in row]
    if cells[0] is None: break
    (name, area, plan) = cells
    if plan in plan_dict:
        temp = plan_dict[plan]
        temp.append(cells)
    else:
        plan_dict[plan] = [cells]
    
for plan, values in plan_dict.items():
    sheet = book.create_sheet(title=plan+'プラン')
    sheet.append(['名前', '住所', 'プラン'])
    for row in values:
        sheet.append(row)
        
book.save(os.path.dirname(__file__) + '/split_sheet2.xlsx')