# -*- coding: utf-8 -*-

import openpyxl as excel
import os

book = excel.load_workbook(os.path.dirname(__file__) + '/test100.xlsx')
sheet = book.active

print('===< 解法1 >===')
it = sheet.iter_rows(min_row=2, min_col=2, max_row=4, max_col=4)
result = []
for row in it:
    r = []
    for cell in row:
        r.append(cell.value)
    result.append(r)
    
print(result)

print('===< 解法2 >===')
it2 = sheet.iter_rows(min_row=2, min_col=2, max_row=4, max_col=4)
result2 = [[cell.value for cell in row] for row in it2]
print(result2)
