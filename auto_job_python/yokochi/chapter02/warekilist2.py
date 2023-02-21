# -*- coding: utf-8 -*-
import openpyxl as excel
import os

book = excel.Workbook()
sheet = book.active

sheet['A1'] = '西暦'
sheet['B1'] = '和暦'

start_year = 1930

for i in range(100):
    seireki = str(start_year + i)
    wareki = '=TEXT("{}/1/1", "ggge年")'.format(seireki)
    
    sheet.cell(row=(2+i), column=1, value=seireki+'年')
    sheet.cell(row=(2+i), column=2, value=wareki)
    
book.save(os.path.dirname(__file__) + '/wareki2.xlsx')