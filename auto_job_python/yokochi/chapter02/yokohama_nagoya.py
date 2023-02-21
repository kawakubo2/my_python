import openpyxl as excel
import os

book = excel.load_workbook(os.path.dirname(__file__) + '/all-customer.xlsx')
sheet = book['名簿']
customers = [['名前', '住所', '購入プラン']]

for row in sheet.iter_rows(min_row=3):
    values = [v.value for v in row]
    (name, address, plan) = values
    if name is None: break
    if address in ('横浜市', '名古屋市'):
        customers.append(values)
        print(values)

new_book = excel.Workbook()
new_sheet = new_book.active
new_sheet['A1'] = '横浜と名古屋の顧客名簿'
for row, row_val in enumerate(customers):
    for col, val in enumerate(row_val):
        c = new_sheet.cell(row+2, col+1)
        c.value = val

new_book.save(os.path.dirname(__file__) + '/yokohama_nagoya.xlsx')