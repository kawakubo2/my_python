import os, requests, urllib, time, re
import openpyxl as excel
from selenium import webdriver
from selenium.webdriver.support.select import Select

#正規表現をあらかじめコンパイルしておく
departure_arrival_pattern = re.compile(r"\s*(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})\s*")
flight_number_pattern = re.compile(r"\s*([a-zA-Z0-9]+)\s*")
price_pattern = re.compile(r"\s*(\d{1,3}),(\d{3})円")
driver = None

def main(target_date="20230115", departure="FUK", arrival="HND"):
    global driver 

    book = excel.Workbook();
    
    get_price_page(target_date, departure, arrival)
    time.sleep(3)
    
    base_dir = os.path.dirname(__file__)
    file_dir = os.path.join(base_dir, "files")
    save_file = os.path.join(file_dir, "price.xlsx")


    ids = {"availabilityResultMainFare": "buttonMainFare", "availabilityResultValueFare": "buttonValueFare", \
        "availabilityResultSuperValueFare": "buttonSuperValueFare", "availabilityResultOtherFare": "buttonOtherFare"}
    tabs = {"availabilityResultMainFare": "主な運賃", "availabilityResultValueFare": "バリュー", \
        "availabilityResultSuperValueFare": "スーパーバリュー", "availabilityResultOtherFare": "その他の運賃"}

    for id, button in ids.items():
        print(f"-----< {tabs[id]} >---")
        fare_button = driver.find_elements_by_css_selector(f"#{button} > a")
        fare_button[0].click()
        time.sleep(2)

        # 新規シートの作成
        sheet = book.create_sheet(title=tabs[id])
         
        # Excel見出し行作成
        excel_headers = ['出発時刻', '到着時刻', '便名', 'プレミアムクラス'];
        head_tds = driver.find_elements_by_css_selector(f"#{id} > table > thead > tr > td")
        print(f"見出しのセル数: {len(head_tds)}")
        for i in range(3, len(head_tds) + 1):
            head_text = driver.find_elements_by_css_selector(f"#{id} > table > thead > tr > td:nth-child({i}) > div > a")
            if head_text:
                excel_headers.append(head_text[0].text.replace("\n", ""))
        print(excel_headers)
        for i in range(1, len(excel_headers) + 1):
            c = sheet.cell(1, i)
            c.value = excel_headers[i - 1]
        
        # Excelデータ行の作成
        #発着時刻はMailFareのみ
        trs = driver.find_elements_by_css_selector(f"#{id} > table > tbody > tr")
        tr_len = len(trs)
        print(f"行数:{tr_len}")
        for i in range(1, tr_len +1):
            #if id == "availabilityResultValueFare":
            departure_arrival = driver.find_elements_by_css_selector(f"#{id} > table > tbody > tr:nth-child({i}) > td.headCell > p.availabilityResultFlightTime")
            departure, arrival = parse_departure_arrival(departure_arrival[0].text)
            print("-----------------")
            # 出発時刻
            c = sheet.cell(i + 1, 1)
            c.value = departure
            # 到着時刻
            c = sheet.cell(i + 1, 2)
            c.value = arrival
            flight_number = driver.find_elements_by_css_selector(f"#{id} > table > tbody > tr:nth-child({i}) > td.headCell > p.availabilityResultFlightDetail > span:nth-child(1)")
            # 便名
            c = sheet.cell(i + 1, 3)
            c.value = flight_number_pattern.search(flight_number[0].text).group(1)
            
            #premiumクラスの料金取得
            premium_class = driver.find_elements_by_css_selector( f"#{id} > table > tbody > tr:nth-child({i}) > td[class='availabilityResultPremiumClass'] div > a")
            c = sheet.cell(i + 1, 4)
            if premium_class:
                # premium
                c.value = parse_price(premium_class[0].text)
            else:
                c.value = '---'
                    
            #premiumクラス以外の料金取得
            tds = driver.find_elements_by_css_selector(f"#{id} > table > tbody > tr:nth-child({i}) > td")
            td_len = len(tds)
            print(F"価格カラム数: {td_len}")
            for j in range(3, td_len +1):
                # 4 - 2するわけ
                # 4は出発時刻、到着時刻、便名、プレミアム。4つのExcelのセルを使っている
                # 2は1番目のtdには出発時刻、到着時刻、便名。2番目にはプレミアムの値が入っている
                # したがって、繰り返し処理するtdタグは3番目から存在するため + 2する必要がある。
                c = sheet.cell(i + 1, j + (4 - 2))
                waiting = driver.find_elements_by_css_selector(f"#{id} > table > tbody > tr:nth-child({i}) > td:nth-child({j}) > div > span:nth-child(1) > label:nth-child(1) > span[class='hasSeatWait'] ")
                if waiting:
                    print(waiting[0].text, end="")
                price = driver.find_elements_by_css_selector(f"#{id} > table > tbody > tr:nth-child({i}) > td:nth-child({j}) > div > span:nth-child(1) > label:nth-child(1) > em ")
                if price:
                    print(f"価格: {price[0].text}")
                    if waiting:
                        c.value = waiting[0].text + ":" + parse_price(price[0].text)
                    else:
                        c.value = parse_price(price[0].text)
                else:
                    full_occupancy = driver.find_elements_by_css_selector(f"#{id} > table > tbody > tr:nth-child({i}) > td:nth-child({j}) > div > span:nth-child(1)")
                    if full_occupancy:
                        # 満席
                        c.value = full_occupancy[0].text

    book.save(save_file)
    driver.close()
    
def get_price_page(target_date="20230115", departure="FUK", arrival="HND"):
    """
    出発地、到着地、搭乗日で検索し、表示されたURLを返す
    """
    global driver

    url = "https://www.ana.co.jp/ja/jp/?type=d"
    driver = webdriver.Chrome()
    driver.get(url)
    time.sleep(5)
    # print(driver.get_attribute("innerHTML"))

    #選択画面の表示
    get_search_form()
    time.sleep(5)
    #出発地の選択
    select_departure(departure)
    #到着地の選択
    select_arrival(arrival)
    #搭乗日の選択
    select_date(target_date)
    time.sleep(5)

def get_search_form():
    """
    検索画面の表示
    """
    global driver
    ticket_submit_button = driver.find_elements_by_css_selector(".be-domestic-reserve-ticket-submit__button")
    ticket_submit_button[0].click()

def select_departure(departure):
    """
    出発地の選択
    """
    global driver
    departure_elem = driver.find_element_by_id("departureAirport")
    select_departure = Select(departure_elem)
    select_departure.select_by_value(departure)

def select_arrival(arrival):
    """
    到着地の選択
    """
    global driver
    arrival_elem = driver.find_element_by_id("arrivalAirport")
    select_arrival = Select(arrival_elem)
    select_arrival.select_by_value(arrival)

def select_date(target_date):
    """
    搭乗日をクリックすると日付を選択できるカレンダーが表示される
    クリックしないと日付が選択できない（日付はreadonlyになっている）
    """
    global driver
    outwardEmbarkationDate_elem = driver.find_element_by_id("outwardEmbarkationDate")
    outwardEmbarkationDate_elem.click()
    time.sleep(3)
    #カレンダーから日付を選択    
    calendar_elem = driver.find_elements_by_css_selector(f"#modal_scroller_calendar_outwardEmbarkationDate td[data-date = '{target_date}'] > a")
    calendar_elem[0].click()
    
    search_btn = driver.find_element_by_xpath('//input[@type="submit"][@value="検索する"]')
    search_btn.click()

def parse_departure_arrival(departure_arrival):
    global departure_arrival_pattern
    result = departure_arrival_pattern.search(departure_arrival)
    if result:
        return(result.group(1), result.group(2))
    else:
        return None
    
def parse_price(price_str):
    global price_pattern
    result = price_pattern.search(price_str)
    if result:
        return result.group(1) + result.group(2)
    else:
        return None

if __name__ == "__main__":
    main()