# coding: UTF-8

import os
import requests
import urllib
import datetime
import time
import re
import sys
import openpyxl as excel
from selenium import webdriver
from selenium.webdriver.support.select import Select

# 正規表現をあらかじめコンパイルしておく
departure_arrival_pattern = re.compile(
    r"\s*(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})\s*")
flight_number_pattern = re.compile(r"\s*([a-zA-Z0-9]+)\s*")
price_pattern = re.compile(r"\s*(\d{1,3}),(\d{3})円")
date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")

ids = {"availabilityResultMainFare": "buttonMainFare", "availabilityResultValueFare": "buttonValueFare",
       "availabilityResultOtherFare": "buttonOtherFare"}
# ids = {"availabilityResultMainFare": "buttonMainFare", "availabilityResultValueFare": "buttonValueFare", \
#    "availabilityResultSuperValueFare": "buttonSuperValueFare", "availabilityResultOtherFare": "buttonOtherFare"}
tabs = {"availabilityResultMainFare": "主な運賃", "availabilityResultValueFare": "バリュー",
        "availabilityResultOtherFare": "その他の運賃"}
# tabs = {"availabilityResultMainFare": "主な運賃", "availabilityResultValueFare": "バリュー", \
#    "availabilityResultSuperValueFare": "スーパーバリュー", "availabilityResultOtherFare": "その他の運賃"}

# 期間の料金を格納するための辞書
price_dict = {}

driver = None

def main(departure="FUK", arrival="HND"):
    global driver

    dates = get_period_dates()
    date_strs = transform_dates(dates)
    print(date_strs)

    for target_date in dates:
        print(f"日付: {target_date}")
        url = "https://www.ana.co.jp/ja/jp/?type=d"
        driver = webdriver.Chrome()
        driver.get(url)
        time.sleep(5)

        get_price_page(target_date, departure, arrival)
        date_str = f"{target_date[0:4]}-{target_date[4:6]}-{target_date[6:8]}"
        time.sleep(3)

        for id, button in ids.items():
            if id not in price_dict:
                price_dict[id] = {}
            print(f"-----< {tabs[id]} >---")
            fare_button = driver.find_elements_by_css_selector(
                f"#{button} > a")
            fare_button[0].click()
            time.sleep(2)

            # Excel見出し行作成
            excel_headers = ['出発時刻', '到着時刻', '便名', 'プレミアムクラス']
            head_tds = driver.find_elements_by_css_selector(
                f"#{id} > table > thead > tr > td")
            print(f"見出しのセル数: {len(head_tds)}")
            for i in range(3, len(head_tds) + 1):
                head_text = driver.find_elements_by_css_selector(
                    f"#{id} > table > thead > tr > td:nth-child({i}) > div > a")
                if head_text:
                    excel_headers.append(head_text[0].text.replace("\n", ""))

            print('excel_headers =', end='')
            print(excel_headers)

            # Excelデータ行の作成
            trs = driver.find_elements_by_css_selector(
                f"#{id} > table > tbody > tr")
            tr_len = len(trs)
            print(f"行数:{tr_len}")
            for i in range(1, tr_len + 1):
                departure_arrival = driver.find_elements_by_css_selector(
                    f"#{id} > table > tbody > tr:nth-child({i}) > td.headCell > p.availabilityResultFlightTime")
                departure_time, arrival_time = parse_departure_arrival(
                    departure_arrival[0].text)
                print("-----------------")
                # 便名

                flight_number = driver.find_elements_by_css_selector(
                    f"#{id} > table > tbody > tr:nth-child({i}) > td.headCell > p.availabilityResultFlightDetail > span:nth-child(1)")
                flight_number_name = flight_number_pattern.search(
                    flight_number[0].text).group(1)

                if flight_number_name in price_dict[id]:
                    print("フライト便名が存在する場合")
                    flight_number_dict = price_dict[id][flight_number_name]
                else:  # 便の最初の処理
                    print("フライト便名が存在しない場合")
                    flight_number_dict = price_dict[id][flight_number_name] = {
                    }
                    flight_number_dict["出発時刻"] = departure_time
                    flight_number_dict["到着時刻"] = arrival_time
                    flight_number_dict["料金表"] = {}
                price_table = flight_number_dict["料金表"]
                if "プレミアム" not in price_table:
                    price_table["プレミアム"] = {}
                for header in excel_headers[4:]:
                    if header not in price_table:
                        price_table[header] = {}
                print("---< price_table >---")
                print(price_table)

                # premiumクラスの料金取得
                premium_class = driver.find_elements_by_css_selector(
                    f"#{id} > table > tbody > tr:nth-child({i}) > td[class='availabilityResultPremiumClass'] div > a")
                if premium_class:
                    # premium
                    price_table["プレミアム"][date_str] = parse_price(
                        premium_class[0].text)
                else:
                    pass

                # premiumクラス以外の料金取得
                tds = driver.find_elements_by_css_selector(
                    f"#{id} > table > tbody > tr:nth-child({i}) > td")
                td_len = len(tds)
                print(F"価格カラム数: {td_len}")
                for j in range(3, td_len + 1):
                    # 4 - 2するわけ
                    # 4は出発時刻、到着時刻、便名、プレミアム。4つのExcelのセルを使っている
                    # 2は1番目のtdには出発時刻、到着時刻、便名。2番目にはプレミアムの値が入っている
                    # したがって、繰り返し処理するtdタグは3番目から存在するため + 2する必要がある。
                    waiting = driver.find_elements_by_css_selector(
                        f"#{id} > table > tbody > tr:nth-child({i}) > td:nth-child({j}) > div > span:nth-child(1) > label:nth-child(1) > span[class='hasSeatWait'] ")
                    if waiting:
                        print(waiting[0].text, end="")
                    price = driver.find_elements_by_css_selector(
                        f"#{id} > table > tbody > tr:nth-child({i}) > td:nth-child({j}) > div > span:nth-child(1) > label:nth-child(1) > em ")
                    if price:
                        print(f"価格: {price[0].text}")
                        # if date_str not in price_table[excel_headers[j + 1]]:
                        #     price_table[excel_headers[j + 1]] = {}
                        if waiting:
                            # 2022-12-14授業後変更分 j - 3 を j + 1
                            price_table[excel_headers[j + 1]][date_str] \
                                = waiting[0].text + ":" + parse_price(price[0].text)
                        else:
                            # 2022-12-14授業後変更分 j - 3 を j + 1
                            price_table[excel_headers[j + 1]
                                        ][date_str] = parse_price(price[0].text)
                    else:
                        full_occupancy = driver.find_elements_by_css_selector(
                            f"#{id} > table > tbody > tr:nth-child({i}) > td:nth-child({j}) > div > span:nth-child(1)")
                        if full_occupancy:
                            # 満席
                            price_table[excel_headers[j + 1]][date_str] = full_occupancy[0].text
        driver.close()
    print("---< 最終結果 >---")
    print(price_dict)
    headers = create_excel_headers(price_dict, date_strs)
    print(headers)
    write_excel(headers, price_dict, date_strs)

def create_excel_headers(price_dict, date_strs):
    header_dict = {}
    for id, fare in price_dict.items():
        header_dict[id] = {} 

        for flight_number, main_prices in fare.items():
            break

        header_dict[id]["flight_number"] = "便名";
        header_dict[id]["depature"] = "出発時刻"
        header_dict[id]["arrival"] = "到着時刻" 
        header_dict[id]["price_type"] = {}
        for price_type, _ in main_prices["料金表"].items():
            header_dict[id]["price_type"][price_type] = date_strs

    return header_dict

def write_excel(headers, price_dict, date_strs):
    book = excel.Workbook()
    base_dir = os.path.dirname(__file__)
    file_dir = os.path.join(base_dir, "files")
    save_file = os.path.join(file_dir, "price.xlsx")

    for id, fare in price_dict.items():
        # 新規シートの作成
        #########################################
        # ヘッダ出力
        #########################################
        sheet = book.create_sheet(title=tabs[id])
        row = 1
        col = 1
        header = headers[id]
        c = sheet.cell(row, col)
        c.value = header['flight_number']
        col += 1
        c = sheet.cell(row, col)
        c.value = header['depature']
        col += 1
        c = sheet.cell(row, col)
        c.value = header['arrival']
        price_type = header['price_type']
        col += 1
        for type, flight_dates in price_type.items():
            c = sheet.cell(row, col)
            c.value = type
            for flight_date in flight_dates:
                c = sheet.cell(row + 1, col)
                c.value = flight_date
                col += 1

        #########################################
        # データ出力
        #########################################
        for flight_number, main_prices in fare.items():
            c = sheet.cell(row + 2, 1)  # 便名
            c.value = flight_number

            c = sheet.cell(row + 2, 2)  # 出発時刻
            c.value = main_prices['出発時刻']

            c = sheet.cell(row + 2, 3)  # 到着時刻
            c.value = main_prices['到着時刻']

            # price_typeとは「プレミアム」、「フレックス」、「バリュー」等のことを指す
            col=4
            for _, date_price_dict in main_prices['料金表'].items():
                for date_str in date_strs:
                    c = sheet.cell(row + 2, col);
                    if date_str in date_price_dict:
                        c.value = date_price_dict[date_str] 
                    else:
                        c.value = ""  
                    col += 1

            row += 1
    book.remove(book['Sheet'])
    book.save(save_file)


def get_price_page(target_date, departure="FUK", arrival="HND"):
    """
    出発地、到着地、搭乗日で検索し、表示されたURLを返す
    """

    # 選択画面の表示
    get_search_form()
    time.sleep(5)
    # 出発地の選択
    select_departure(departure)
    # 到着地の選択
    select_arrival(arrival)
    # 搭乗日の選択
    select_date(target_date)
    time.sleep(5)


def input_date(prompt):
    while True:
        date = input(prompt + ": ")
        if date_pattern.search(date):
            return date
        print("日付形式に誤りがあります")
        print("入力形式: 年4桁-月2桁-日2桁")
        print("入力例: 2025-02-03")


def get_period_dates():
    print('--- 期間の入力 ---')
    start_date = input_date('開始日付')
    end_date = input_date('終了日付')
    start_date_obj = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    end_date_obj = datetime.datetime.strptime(end_date, "%Y-%m-%d")
    if (start_date_obj > end_date_obj):
        print("開始日付と終了日付が逆転しています。再度入力してください。")
        sys.exit(1)

    dates = [start_date_obj.strftime('%Y%m%d')]
    while (start_date_obj < end_date_obj):
        start_date_obj = start_date_obj + datetime.timedelta(days=1)
        dates.append(start_date_obj.strftime('%Y%m%d'))
    print(dates)
    return dates

def transform_dates(dates):
    result = []
    for d in dates:
        date_str = f"{d[0:4]}-{d[4:6]}-{d[6:8]}"
        result.append(date_str)
    return result

def get_search_form():
    """
    検索画面の表示
    """
    ticket_submit_button = driver.find_elements_by_css_selector(
        ".be-domestic-reserve-ticket-submit__button")
    ticket_submit_button[0].click()


def select_departure(departure):
    """
    出発地の選択
    """
    departure_elem = driver.find_element_by_id("departureAirport")
    select_departure = Select(departure_elem)
    select_departure.select_by_value(departure)


def select_arrival(arrival):
    """
    到着地の選択
    """
    arrival_elem = driver.find_element_by_id("arrivalAirport")
    select_arrival = Select(arrival_elem)
    select_arrival.select_by_value(arrival)


def select_date(target_date):
    """
    搭乗日をクリックすると日付を選択できるカレンダーが表示される
    クリックしないと日付が選択できない（日付はreadonlyになっている）
    """
    outwardEmbarkationDate_elem = driver.find_element_by_id(
        "outwardEmbarkationDate")
    outwardEmbarkationDate_elem.click()
    time.sleep(3)
    # カレンダーから日付を選択
    calendar_elem = driver.find_elements_by_css_selector(
        f"#modal_scroller_calendar_outwardEmbarkationDate td[data-date = '{target_date}'] > a")
    calendar_elem[0].click()

    search_btn = driver.find_element_by_xpath(
        '//input[@type="submit"][@value="検索する"]')
    search_btn.click()


def parse_departure_arrival(departure_arrival):
    global departure_arrival_pattern
    result = departure_arrival_pattern.search(departure_arrival)
    if result:
        return (result.group(1), result.group(2))
    else:
        return None


def parse_price(price_str):
    global price_pattern
    result = price_pattern.search(price_str)
    if result:
        return result.group(1) + result.group(2)
    else:
        return None


if __name__ == "__main__":
    main()
