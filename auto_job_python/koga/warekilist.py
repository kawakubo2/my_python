import openpyxl as excel
import os

wareki_table = [
    {"name": "明治", "start": 1868, "end": 1912},
    {"name": "大正", "start": 1912, "end": 1926},
    {"name": "昭和", "start": 1926, "end": 1989},
    {"name": "平成", "start": 1989, "end": 2019},
    {"name": "令和", "start": 2019, "end": 9999}
]

def seireki_wareki(year):
    for w in wareki_table:
        if w["start"] <= year < w["end"]:
            y = str(year - w["start"] + 1) + "年"
            if y == "1年": y = "元年"
            return w["name"] + y
    return "不明"

book = excel.Workbook()
sheet = book.active
# シートのヘッダに説明を入れる
sheet["A1"] = "西暦"
sheet["B1"] = "和暦"

start_y = 1930
for i in range(100):
    sei = start_y + i
    wa = seireki_wareki(sei)
    # シートに設定
    sheet.cell(row=(i+2), column=1, value=str(sei) + "年")
    sheet.cell(row=(i+2), column=2, value=str(wa))

book.save(os.path.dirname(__file__) + "/wareki.xlsx")
