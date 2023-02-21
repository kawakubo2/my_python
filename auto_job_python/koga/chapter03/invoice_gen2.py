# テンプレートから請求書を作る
# ファイル名を指定
import openpyxl as excel
import os


template_file = os.path.dirname(__file__) + '/invoice-template.xlsx'
save_dirname  = os.path.dirname(__file__)

# 設定するデータ
# name = '田中一郎'
subject = '1月分の請求書'
# items = [
#     ['リンゴ', 5, 320],
#     ['バナナ', 8, 210],
#     ['メロン', 1, 1200]
# ]

seikyu_data = [
    { "name": "田中一郎", "items": [['リンゴ', 5, 320], ['バナナ', 8, 210], ['メロン', 1, 1200]]},
    { "name": "山田太郎", "items": [['オレンジ', 10, 200], ['バナナ', 4, 210], ['レモン', 20, 100]]},
    { "name": "横山花子", "items": [['グレープ', 8, 150], ['メロン', 5, 1200]]},
]

for data in seikyu_data:
    # テンプレートを開く
    book = excel.load_workbook(template_file)
    sheet = book.active
    # テンプレートに名前と件名を書き込む
    sheet["B4"] = data["name"] 
    sheet["C10"] = subject
    # 内訳を連続で書き込む
    total = 0
    for i, it in enumerate(data["items"]):
        summary, count, price = it
        subtotal = count * price
        total += subtotal
        # シートに書き込む
        row = 15 + i
        sheet.cell(row, 2, summary)
        sheet.cell(row, 5, count)
        sheet.cell(row, 6, price)
        sheet.cell(row, 7, subtotal)
    # 請求金額(合計金額)を書き込む
    sheet["C11"] = total
    # ブックを保存
    book.save(save_dirname + "/" + data["name"] + ".xlsx")