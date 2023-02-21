import openpyxl as excel
import json
import os

# 各種設定
CURRENT = os.path.dirname(__file__)
in_file = CURRENT + '/matome.json'
out_dir = CURRENT + '/invoice02'
template_file = CURRENT + '/invoice-template.xlsx'

subject = "2月分の請求書"

# メイン処理
def gen_invoice():
    # JSONファイルを読み込む
    with open(in_file, "rt") as fp:
        users = json.load(fp)
    # 顧客ごとに請求書を作成
    for name, data in users.items():
        make_user_invoice(name, data)

# テンプレートにデータを流し込む
def make_user_invoice(name, data):
    # テンプレートを読み込む
    book = excel.load_workbook(template_file)

    sheet = book.active
    # シートに名前と件名と金額を書き込む
    sheet["B4"] = name
    sheet["C10"] = subject
    sheet["C11"] = data["total"]
    # 内訳を連続で書き込む
    for i, it in enumerate(data["items"]):
        date, summary, cnt, price = it
        row = 15 + i
        sheet.cell(row, 2, summary + '(' + date + ')')
        sheet.cell(row, 5, cnt)
        sheet.cell(row, 6, price)
        sheet.cell(row, 7, cnt*price)
    # 請求書ファイルに保存
    out_file = out_dir + '/' + name + '様.xlsx'
    book.save(out_file)
    print("save: ", out_file)

if __name__ == "__main__":
    gen_invoice()