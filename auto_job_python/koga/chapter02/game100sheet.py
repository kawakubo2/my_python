import openpyxl as excel
import os, random

atari = random.randint(1, 100)

book = excel.Workbook()
book.active["B2"] = "当たりが書かれたシートを探そう"

for i in range(1, 101):
    sname = str(i) + "番"
    sheet = book.create_sheet(title = sname)
    word = "ハズレ"
    if i == atari: word = "当たり"
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y + 1, x + 1)
            c.value = word

book.save(os.path.dirname(__file__) + "/game100.xlsx")
print("ok, atari=", atari)