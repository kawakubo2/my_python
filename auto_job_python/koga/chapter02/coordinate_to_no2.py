import openpyxl as excel

book = excel.Workbook()
sheet = book.active

# セル名から行番号・列番号へ変換
cell = sheet["C2"]
(row, col) = (cell.row, cell.column)
print("C2=({},{})".format(row,col))

# 行番号・列番号から列名へ変換
cell = sheet.cell(row=2, column=3)
cdt = cell.coordinate
print("(2,3)={}".format(cdt))

def get_cell_name(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    return cell.coordinate

for y in range(2, 5):
    for x in range(2, 5):
        print("{} ".format(get_cell_name(sheet, y, x)), end="")
    print()