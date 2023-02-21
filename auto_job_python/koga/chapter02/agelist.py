import openpyxl as excel
import datetime
import os

book = excel.Workbook()
sheet = book.active

# 今年を得る
thisyear = datetime.date.today().year

for i in range(80):
    age = i
    year = thisyear - i
    age_cell = sheet.cell(i+1, 1)
    age_cell.value = str(i) + '才'
    year_cell = sheet.cell(i+1, 2)
    year_cell.value = str(year) + '年'

book.save(os.path.dirname(__file__) + '/agelist.xlsx')

print(os.path.abspath(__file__))