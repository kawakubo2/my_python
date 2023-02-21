import openpyxl as excel
import os

book = excel.Workbook()
sheet = book.active

# セル名を指定する
sheet['A1'] = '勤勉な人の計画は必ず成功する'

# 行番号、列番号、値を指定する
sheet.cell(row=2, column=1, value="猿の尻笑い")

# 行番号、列番号を指定してセルを取り出し、valueプロパティに値を設定
cell = sheet.cell(row=3, column=1)
cell.value = '探すのに時があり諦めるに時がある'

book.save(os.path.dirname(__file__) + '/cell_w.xlsx')