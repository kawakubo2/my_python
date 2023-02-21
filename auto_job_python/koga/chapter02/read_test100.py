import openpyxl as excel
import os

book = excel.load_workbook(os.path.dirname(__file__) + "/test100.xlsx")
sheet = book.active

# セル名で取得
print(sheet["H2"].value)

# 行番号、列番号でセルを取り出し、valueでそのセルの値を取り出す
cell = sheet.cell(row=2, column=8)
print(cell.value)

# 行番号、列番号でセルを取り出し、valueで値を取り出す
print(sheet.cell(row=2, column=8).value)
