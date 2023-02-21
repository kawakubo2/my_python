import openpyxl as excel
import os
book = excel.load_workbook(os.path.dirname(__file__) + '/all-customer.xlsx')
sheet = book["名簿"]

for row in sheet.iter_rows(min_row=3):
    cells = [v.value for v in row]
    if cells[0] is None: break
    print(cells)
    name, area, plan = cells
    sname = plan + "プラン"
    if sname not in book.sheetnames:
        to_sheet = book.create_sheet(title=sname)
        to_sheet.append(["名前", "住所", "プラン"])
    else:
        to_sheet = book[sname]
    to_sheet.append(cells)

book.save(os.path.dirname(__file__) + "/split_sheet.xlsx")