import openpyxl as excel
import os

book = excel.Workbook()
sheet = book.active

days = ['2021/3/3', '2021/3/4', '2021/3/5']
for i, day in enumerate(days):
    sheet.cell(row=(i+1), column=1, value=day)
    sheet.cell(row=(i+1), column=2, value="{}".formt('=TEXT("A"' + str(i+1) + ',"ggge年m月d日"')

book.save(os.path.dirname(__file__) + "/susiki.xlsx")
