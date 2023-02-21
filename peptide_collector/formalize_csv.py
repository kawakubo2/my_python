import openpyxl as excel
import os

BASE_DIR   = 'c:/temp/kittaka_data'
# 編集前のCSVファイルの格納場所
SOURCE_DIR = os.path.join(BASE_DIR, 'source')
# 編集後のCSVファイルの格納場所
# 演習後のCSVファイルがアップロードするファイルとなるので、後続の工程の
# 定数と合わせた。
INPUT_DIR = os.path.join(BASE_DIR, 'input')

def main():
    book_path = os.path.join(SOURCE_DIR, 'タンパク質定量値Excelデータ.xlsx')
    book = excel.load_workbook(book_path)
    new_book = excel.Workbook()
    sheet = book.copy_worksheet(book.worksheets[0])
    new_sheet = new_book.worksheets[0]

    for row in sheet:
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value

    new_sheet.delete_rows(1, 1)
    new_sheet.delete_cols(1, 6)
    new_sheet.delete_cols(2, 5)
    if os.path.exists(book_path):
        os.remove(book_path)
    # new_book.save(book_path)

    # 行列の入替
    max_row = new_sheet.max_row
    max_col = new_sheet.max_column
    
    new_new_sheet = book.create_heet()
    
    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            c1 = new_sheet.cell(row=i, column=j)
            c2 = new_new_sheet.cell(row=j, column=i)
            c2.value = c1.value
    
    book.save(os.join.path(INPUT_DIR, 'タンパク質定量値Excelデータ-編集後.xlsx'))
    

if __name__ == '__main__':
    main()

